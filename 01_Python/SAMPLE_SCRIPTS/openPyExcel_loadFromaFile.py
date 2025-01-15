# OPEN PY EXCEL 
# Loading from a File 

# We can use openpyxl.load_workbook() to open an existing workbook 

from openpyxl import load_workbook 

wb = load_workbook('Dmart_Purchase.xlsx') # "C:\Users\Sayed\Desktop\PythonRoadMap Dec 2022\PythonClass_Dec2022\Dmart_Purchase.xlsx"
print(wb.sheetnames)

sheet = wb.get_sheet_by_name("Purchase")   # wb.active 
print(sheet)
print(sheet['D1'].value)

cv = sheet.cell(row=2, column=4) # to read a particular cell value 

print(f"Name of the Item: {cv.value}")

# we can loop through worksheets 

for sheet in wb:
	print(sheet.title)

#
for row in sheet.values:
	for value in row:
		print(value)



