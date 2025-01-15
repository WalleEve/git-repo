#Openpyexcel:
import openpyxl

file_path = "D:\\Class\\TKINTER_PROJECTS\\BankTransactions\\dataFile.xlsx"


wb = openpyxl.load_workbook(file_path)
print(type(wb)) # View the type of wb [object]

sheets = wb.sheetnames # To view total number  of sheets available
print(sheets)
print(wb.active.title) # show the current active sheet

sh1 = wb['cust_dtl'] # fetch perticular sheet
print(type(sh1))

# Get the data from Excel shell wise
# Option 1:
data = sh1["B2"].value # Get the value from shell B2
print(data)

data1 = wb["cust_dtl"]["A2"].value # We can fetch the data from a perticular shell
print(data1)

# option 2:
data3 = sh1.cell(3, 2).value # row = 3, column = 2
print(data3)

data4 = sh1.cell(4, 2).value  # row = 4, column = 2
print(data4)

data5 = sh1.cell(row = 4, column=5).value
print(data5)

# print(wb.get_sheet_by_name("cust_dtl").cell(row = 4, column = 5).value)
#wb.get_sheet_by_name("cust_dtl").cell(row = 4, column = 5).value

# To View number of rows and columns in a perticular sheet
sh1 = wb["cust_dtl"]
row = sh1.max_row
column = sh1.max_column
print(row)
print(column)


for i in range(1, row + 1 ):
    for j in range(1, column + 1):
        print(sh1.cell(i, j).value)


sh1.cell(row = 4, column = 4, value = 1000) # Change the value
wb.save("dataFile.xlsx") # Save the updated value
#wb.save("D:\\Class\\TKINTER_PROJECTS\\BankTransactions\\dataFile.xlsx")
