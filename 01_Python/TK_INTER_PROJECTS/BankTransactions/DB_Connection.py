# To install xlrd: pip install xlrd
# Reading an excel file using Python
import openpyxl as xl 

""" id name account_number balane password """	
# Excel File Location 
file_path = "D:\\Class\\TKINTER_PROJECTS\\BankTransactions\\dataFile.xlsx"

wb_obj = xl.load_workbook(file_path)
my_sheet_obj = wb_obj.active 
my_cell_obj = my_sheet_obj.cell(row = 1, column = 1)
my_row = my_sheet_obj.max_row

for i in range(2, my_row + 1):
	cell_obj = my_sheet_obj.cell(row = i, column = 1)
	print(cell_obj.value)
