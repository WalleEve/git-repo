import sys
import openpyxl as xl

class Customer:
	bankname = "ISL Bank"
	bnkaddress = "Ameerpet Hyderabad 500016"
	file_path = "D:\\Class\\TKINTER_PROJECTS\\BankTransactions\\dataFile.xlsx" # FIle path for excel data store
	wb_obj = xl.load_workbook(file_path) # Excel Workbook object
	sheet_obj = wb_obj.active # Active sheet object

	def __init__(self, name, amount = 0.0):
		self.name = name
		self.amount = amount

	def withdraw(self, amt):
		if amt > self.amount:
			print("Insufficent Balance...")
			sys.exit()
		self.amount = self.amount - amt
		print("Balance after withdraw: ", self.amount)

	def deposit(self, amt):
		self.amount = self.amount + amt
		print("Balance after deposit: ", self.amount)

	def acc_create(self):

		mx_row = Customer.sheet_obj.max_row # Maximum number of rows

		id = Customer.sheet_obj.cell(row = mx_row + 1, column = 1, value = mx_row + 1)
		Customer.sheet_obj.cell(row = mx_row + 1, column = 2, value= self.name)
		Customer.sheet_obj.cell(row = mx_row + 1, column = 3, value= mx_row + 1)
		Customer.sheet_obj.cell(row = mx_row + 1, column = 4, value= self.amount)
		Customer.sheet_obj.cell(row = mx_row + 1, column = 5, value= self.name)
		Customer.wb_obj.save("dataFile.xlsx")
		self.account_no = mx_row + 1
		print("Account Created Successfully")
		print("Name: ", self.name)
		print("Account Number: ", id.value)
		print("Balance: ", self.amount)
		print("Transaction Password: ", self.name)

	def login(self, acc_no, password):
		mx_row = Customer.sheet_obj.max_row # Maximum number of rows
		for i in range(2, mx_row + 1):
			name_cell = Customer.sheet_obj.cell(row = i, column = 2)
			acc_cell = Customer.sheet_obj.cell(row = i, column = 3)
			balacne_cell = Customer.sheet_obj.cell(row = i, column = 4)
			password_cell = Customer.sheet_obj.cell(row = i, column = 5)
			if acc_cell.value == account_no and password_cell.value == password:
				flag = True
				break
			else:
				flag = False
		return flag













print("Welcome To ISL BANK")
print("Ameerpet Hyderabad 500016")
print("*" * 50, "\n")


# Main Operational Code:
while True:
	print("Please choose option: \n n - New Account\n o - Old Account")
	flag = input("> ")

	if flag.lower() == "n":
		print("Opening New Account")
		name = input("Please enter name: ")
		na = Customer(name)
		na.acc_create()

		sys.exit()
	elif flag.lower() == "o":
		print("Welcome")
		account_no = int(input("Enter your Account number: "))
		password = input("Enter Password: ")
		flag = na.login(account_no, password)
		if flag == True:
			while True:
				print("Please choose option:\n d - Deposit\n w - Withdraw\n e - Exit ")
				flag = input("> ")

				if flag.lower() == "e":
					print("Thank you for Banking with us")
					sys.exit()

				elif flag.lower() == "d":
					amount = float(input("Please enter amount: "))
					c.deposit(amount)

				elif flag.lower() == "w":
					amount = float(input("Please enter amount: "))
					c.withdraw(amount)
				else:
					print("Please choose valid option !")

		else:
			print("Invalid username or password")
			sys.exit()

	else:
		print("Invalid option")
		sys.exit()
